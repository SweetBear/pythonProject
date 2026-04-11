#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2024/7/19 下午3:55
@Author  : Bill Fang
@File    : ExcelDeal.py
@Desc    : 
"""
from openpyxl import Workbook,load_workbook
import mysql.connector
import requests
import os
from PIL import Image 


def check_img_direction(img_path):

    image = Image.open(img_path)
    # 获取exif信息
    exif = image._getexif()

    # 提取旋转角度信息
    if exif:
        orientation = exif.get(274)
        if orientation == 3:
            print("图像被逆时针旋转了180度")
        elif orientation == 6:
            print(img_path + "图像被逆时针旋转了90度")
            rotated_image  = image.rotate(360, expand=True)
            #img_path = img_path.replace("新", "生成")
            rotated_image .save(img_path)
        elif orientation == 8:
            print("图像被逆时针旋转了270度")
        else:
            print(img_path + "图像未旋转")
    else:
        print(img_path + "无法获取旋转信息")

def deal_local_img():
    class_dict = {}
    name_dict = {}
    student_excel_path = ("E:\\img_test\\123.xlsx")

    # 加载工作簿
    workbook = load_workbook(filename=student_excel_path)
    # 选择工作表
    sheet = workbook['Sheet1']
    # 读取整个工作表数据
    for row in sheet.iter_rows(min_row=3, max_col=10, values_only=True):
        #print(str(row[1]) + "--->" + str(row[2]) + "--->" + str(row[3]) + "--->" + str(row[6]))
        if str(row[9]) != "" and str(row[9]) != None:

                class_dict[row[1] + row[9]] = row[2]+row[3]
                name_dict[row[1] + row[9]] = row[1]
        else:
            print(str(row[1]) + "身份证不存在")
    print(class_dict)
    print(name_dict)
    print("====================================================================")
    folder_path = "E:\\img_test\\123\\"
    file_names = os.listdir(folder_path)
    # 遍历文件夹内的所有文件
    for i, file_name in enumerate(file_names):

        # 获取文件的扩展名
        ext = os.path.splitext(file_name)[1]
        file_name_str = file_name.split('.')[0]
        if i == 0:
            output_path = 'E:\\img_test\\img_test\\无班级\\'
            if not os.path.exists(output_path):
                os.makedirs(output_path)
        if file_name_str not in name_dict:
            print(file_name_str + "==>" + ext)
            os.rename(os.path.join(folder_path, file_name), os.path.join('E:\\img_test\\img_test\\无班级\\', file_name_str + ext))
        else:
            output_path = 'E:\\img_test\\img_test\\' + class_dict.get(file_name_str) + "\\"
            if not os.path.exists(output_path):
                os.makedirs(output_path)

            try:
                print(os.path.join(folder_path, file_name) + "==>" + os.path.join(output_path, name_dict[file_name_str]  + ext))
                os.rename(os.path.join(folder_path, file_name), os.path.join(output_path, name_dict[file_name_str] + ext))
            except FileExistsError as e:
                #result_file_name = result_file_name.replace(id_dict.get(file_name_str),
                                                            #id_dict.get(file_name_str) + "_1")
                #os.rename(os.path.join(folder_path, file_name), os.path.join(output_path, result_file_name))
                print(e)


def pre_deal():

    folder_path = "E:\\img_test\\123\\"
    file_names = os.listdir(folder_path)
    # 遍历文件夹内的所有文件
    for i, file_name in enumerate(file_names):
        # 获取文件的扩展名
        ext = os.path.splitext(file_name)[1]
        file_name_str = file_name.split('.')[0]
        if ' ' in file_name_str:
            print(file_name_str + "==============")
            file_name_str = file_name_str.replace(' ', '')
            os.rename(os.path.join(folder_path, file_name), os.path.join(folder_path, file_name_str + ext))

if __name__ == '__main__':
    #pre_deal();
    deal_local_img()
    #folder_path = "E:\\img_test\\新"
    #file_names = os.listdir(folder_path)
    # 遍历文件夹内的所有文件
    #img_url = 'E:\\img_test\\新\\320724200904090011.jpg'
    #check_img_direction(img_url)
    #for i, file_name in enumerate(file_names):
        #img_url = folder_path + '\\' + file_name
        #check_img_direction(img_url)










