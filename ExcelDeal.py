#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2024/7/19 下午3:55
@Author  : Bill Fang
@File    : ExcelDeal.py
@Desc    : 
"""
from openpyxl import Workbook,load_workbook
import mysql.connector
import requests
import re

import uuid


def dealPhoneStr(dpt_name, per_name, phones):
    db_str = ""
    if "/" in phones:
        tmp = ""
        phoneItem = phones.split("/")
        for item in phoneItem:
            if tmp != "":
                tmp = tmp + "/" + item.replace(item, '<a href="tel:' + item + '">' + item + '</a>')
            else:
                tmp = item.replace(item, '<a href="tel:' + item + '">' + item + '</a>')
        tmp = "<div class=\"rbox\">" + tmp + "</div>"
        db_str = "<div class=\"cbox\">" + dpt_name + "，" + per_name +  "，" +  tmp + "</div>"
    else:
        phones = phones.replace(phones, '<a href="tel:' + phones + '">' + phones + '</a>')
        db_str = "<div class=\"rbox\">" + dpt_name + "，" + per_name +  "，" +  phones + "</div>"
    return db_str


def dealHotLine(hotline):
    if hotline == '':
        return ""
    if "\n" in hotline:
        hotline = hotline.replace("\n", "")
    if "；" in hotline:
        db_str = ""
        tmp_str = hotline.split("；")
        for tmp in tmp_str:
            print(tmp)
            items = tmp.split("，")
            dpt_name = items[0]
            per_name = items[1]
            phones = items[2]

            phone_str = dealPhoneStr(dpt_name, per_name, phones)
            db_str = db_str + phone_str
        return db_str
    else:
        print(hotline)
        items = hotline.split("，")
        dpt_name = items[0]
        per_name = items[1]
        phones = items[2]

        phone_str = dealPhoneStr(dpt_name, per_name, phones)


        return phone_str





if __name__ == '__main__':

    # 创建数据库连接
    db = mysql.connector.connect(
        host="",  # MySQL服务器地址
        port="3306",  # MySQL服务器端口
        user="root",  # 用户名
        password="",  # 密码
        database="tw_web"  # 数据库名称
    )

    # 创建游标对象，用于执行SQL查询
    cursor = db.cursor()

    stu_dict = {}
    student_excel_path = "D:\\work\\中型企业贷款.xlsx"

    # 加载工作簿
    workbook = load_workbook(filename=student_excel_path)
    # 选择工作表
    sheet = workbook['Sheet1']
    # 读取整个工作表数据
    # 读取整个工作表数据
    i = 0
    for row in sheet.iter_rows(min_row=4, values_only=True):
        bank_name = str(row[0])
        classify = '3'
        policy_info = str(row[1])
        target_object = str(row[2])
        application_condition = str(row[3])
        time_limit = str(row[4])
        hotline = str(row[5])
        sort = i
        hotline = dealHotLine(hotline)

        uid = uuid.uuid4()
        uid = str(uid).replace('-', '')

        sql = "INSERT INTO `tw_web`.`abc_renewal_loan_info` (`code`, `sort`,`classify`, `bank_name`, `policy_info`, `target_object`, `application_condition`, `time_limit`, `hotline`) VALUES ('" + uid + "', '"  + str(sort) + "', '" + classify + "', '" + bank_name + "', '" + policy_info + "', '" + target_object + "', '" + application_condition + "', '" + time_limit + "','" + hotline + "');"

        print(i)
        i = i + 1
        cursor.execute(sql)

    # 提交更新
    db.commit()




